#!/usr/bin/env python3
"""Build a full-site Excel report of all about.ups.com URLs, one sheet per locale."""
import json
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = "/backups/8497785/ups-unified-site-eds-foundation/repo/catalog-fullsite"
OUT = "/backups/8497785/ups-unified-site-eds-foundation/repo/content/fullsite-urls-report.xlsx"

urls_all = json.load(open(f"{BASE}/urls-all.json"))["analysis-urls-all"]
grouped = json.load(open(f"{BASE}/urls-grouped.json"))["analysis-urls-grouped"]
all_urls = [u["url"] for u in urls_all["urls"]]

# Group URLs by locale (first two path segments)
locale_urls = {}
for u in all_urls:
    parts = [p for p in urlparse(u).path.split("/") if p]
    key = "/".join(parts[:2]) if len(parts) >= 2 else (parts[0] if parts else "/")
    locale_urls.setdefault(key, []).append(u)

def content_area(u, locale_parts):
    parts = [p for p in urlparse(u).path.split("/") if p]
    rest = parts[len(locale_parts):]
    rest = [r for r in rest if not r.replace(".html", "").isdigit()]
    if not rest:
        return "(home)"
    if rest[-1].endswith(".html") and len(rest) > 1:
        rest = rest[:-1]
    return "/" + "/".join(rest[:2])

# Styling
HEADER_FILL = PatternFill("solid", fgColor="351C15")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GOLD_FILL = PatternFill("solid", fgColor="FFC400")
TITLE_FONT = Font(bold=True, size=14, color="242424")
BOLD = Font(bold=True)
LINK = Font(color="0662BB", underline="single")
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------- Summary ----------
ws = wb.active
ws.title = "Summary"
ws["A1"] = "about.ups.com — Full-Site URL Inventory (per locale)"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:D1")
ws["A3"] = f"Total URLs: {len(all_urls):,}"
ws["A4"] = f"Locales: {len(locale_urls)}"
ws["A5"] = "Scope: https://about.ups.com/ (all 28 country sitemaps, no sampling)"
ws["A6"] = "Note: us/en already fully cataloged (18 templates, 693 analyzed pages)"
for r in (3, 4):
    ws[f"A{r}"].font = BOLD

hrow = 8
for c, h in enumerate(["#", "Locale", "URL Count", "Top Content Areas"], 1):
    cell = ws.cell(row=hrow, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER

def safe_sheet_name(name, used):
    n = name.replace("/", "_")[:31]
    base = n
    i = 1
    while n in used:
        n = f"{base[:28]}_{i}"
        i += 1
    used.add(n)
    return n

used_names = set(["Summary", "All URLs"])
sheet_for = {}
for lk in sorted(locale_urls):
    sheet_for[lk] = safe_sheet_name(lk, used_names)

for i, lk in enumerate(sorted(locale_urls), 1):
    lparts = lk.split("/")
    areas = {}
    for u in locale_urls[lk]:
        a = content_area(u, lparts)
        areas[a] = areas.get(a, 0) + 1
    top = ", ".join(f"{k} ({v})" for k, v in sorted(areas.items(), key=lambda x: -x[1])[:5])
    r = hrow + i
    for c, v in enumerate([i, lk, len(locale_urls[lk]), top], 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = WRAP if c == 4 else TOP
    link_cell = ws.cell(row=r, column=2)
    link_cell.hyperlink = f"#'{sheet_for[lk]}'!A1"
    link_cell.font = LINK

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 12
ws.column_dimensions["C"].width = 11
ws.column_dimensions["D"].width = 90
ws.freeze_panes = "A9"

# ---------- All URLs (flat) ----------
allws = wb.create_sheet("All URLs")
for c, h in enumerate(["Locale", "Content Area", "URL", "Path"], 1):
    cell = allws.cell(row=1, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
row = 2
for lk in sorted(locale_urls):
    lparts = lk.split("/")
    for u in sorted(locale_urls[lk]):
        allws.cell(row=row, column=1, value=lk).border = BORDER
        allws.cell(row=row, column=2, value=content_area(u, lparts)).border = BORDER
        uc = allws.cell(row=row, column=3, value=u)
        uc.hyperlink = u
        uc.font = LINK
        uc.border = BORDER
        allws.cell(row=row, column=4, value=urlparse(u).path).border = BORDER
        row += 1
allws.column_dimensions["A"].width = 12
allws.column_dimensions["B"].width = 32
allws.column_dimensions["C"].width = 95
allws.column_dimensions["D"].width = 55
allws.freeze_panes = "A2"
allws.auto_filter.ref = f"A1:D{row-1}"

# ---------- Per-locale sheets ----------
for lk in sorted(locale_urls):
    lparts = lk.split("/")
    s = wb.create_sheet(sheet_for[lk])
    s["A1"] = f"Locale: {lk}"
    s["A1"].font = TITLE_FONT
    s["A1"].fill = GOLD_FILL
    s.merge_cells("A1:C1")
    s["A2"] = "URL count:"
    s["A2"].font = BOLD
    s["B2"] = len(locale_urls[lk])
    hr = 4
    for c, h in enumerate(["#", "Content Area", "URL"], 1):
        cell = s.cell(row=hr, column=c, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
    for i, u in enumerate(sorted(locale_urls[lk]), 1):
        rr = hr + i
        s.cell(row=rr, column=1, value=i).border = BORDER
        s.cell(row=rr, column=2, value=content_area(u, lparts)).border = BORDER
        uc = s.cell(row=rr, column=3, value=u)
        uc.hyperlink = u
        uc.font = LINK
        uc.border = BORDER
    s.column_dimensions["A"].width = 6
    s.column_dimensions["B"].width = 34
    s.column_dimensions["C"].width = 95
    s.freeze_panes = "A5"
    s.auto_filter.ref = f"A{hr}:C{hr + len(locale_urls[lk])}"

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: {len(wb.sheetnames)} (Summary, All URLs, + {len(locale_urls)} locale sheets)")
print(f"Total URLs listed: {sum(len(v) for v in locale_urls.values()):,}")
