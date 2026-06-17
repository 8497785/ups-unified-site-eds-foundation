#!/usr/bin/env python3
"""Build an Excel report of pages grouped by template type."""
import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CATALOG = "/backups/8497785/ups-unified-site-eds-foundation/repo/catalog/template-catalog.json"
OUT = "/backups/8497785/ups-unified-site-eds-foundation/repo/content/template-pages-report.xlsx"

with open(CATALOG) as f:
    catalog = json.load(f)

templates = catalog["templates"]
templates_sorted = sorted(templates, key=lambda t: len(t["urls"]), reverse=True)

# Styling
HEADER_FILL = PatternFill("solid", fgColor="351C15")  # UPS brown
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
GOLD_FILL = PatternFill("solid", fgColor="FFC400")     # UPS gold
TITLE_FONT = Font(bold=True, size=14, color="242424")
BOLD = Font(bold=True)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP = Alignment(vertical="top")
thin = Side(style="thin", color="D0D0D0")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()

# ---------- Sheet 1: Summary ----------
ws = wb.active
ws.title = "Summary"
ws["A1"] = "UPS About Site — Template Catalog Summary"
ws["A1"].font = TITLE_FONT
ws.merge_cells("A1:D1")
ws["A3"] = f"Total templates: {len(templates)}"
ws["A4"] = f"Total pages: {sum(len(t['urls']) for t in templates)}"
ws["A3"].font = BOLD
ws["A4"].font = BOLD

headers = ["#", "Template", "Page Count", "Description"]
hrow = 6
for c, h in enumerate(headers, 1):
    cell = ws.cell(row=hrow, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
    cell.alignment = Alignment(vertical="center")

for i, t in enumerate(templates_sorted, 1):
    r = hrow + i
    vals = [i, t["name"], len(t["urls"]), t.get("description", "")]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=c, value=v)
        cell.border = BORDER
        cell.alignment = WRAP if c == 4 else TOP
    # link to the per-template sheet
    ws.cell(row=r, column=2).hyperlink = f"#'{t['name'][:31]}'!A1"
    ws.cell(row=r, column=2).font = Font(color="0662BB", underline="single")

ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 24
ws.column_dimensions["C"].width = 12
ws.column_dimensions["D"].width = 70
ws.freeze_panes = "A7"

# ---------- Sheet 2: All Pages (flat) ----------
allws = wb.create_sheet("All Pages")
flat_headers = ["Template", "Page Count (template)", "URL", "Path"]
for c, h in enumerate(flat_headers, 1):
    cell = allws.cell(row=1, column=c, value=h)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.border = BORDER
row = 2
for t in templates_sorted:
    for url in t["urls"]:
        path = url.replace("https://about.ups.com", "")
        allws.cell(row=row, column=1, value=t["name"]).border = BORDER
        allws.cell(row=row, column=2, value=len(t["urls"])).border = BORDER
        uc = allws.cell(row=row, column=3, value=url)
        uc.hyperlink = url
        uc.font = Font(color="0662BB", underline="single")
        uc.border = BORDER
        allws.cell(row=row, column=4, value=path).border = BORDER
        row += 1
allws.column_dimensions["A"].width = 22
allws.column_dimensions["B"].width = 20
allws.column_dimensions["C"].width = 90
allws.column_dimensions["D"].width = 60
allws.freeze_panes = "A2"
allws.auto_filter.ref = f"A1:D{row-1}"

# ---------- Per-template sheets ----------
used_names = set()
for t in templates_sorted:
    base = t["name"][:31]
    name = base
    n = 1
    while name in used_names:
        name = f"{base[:28]}_{n}"
        n += 1
    used_names.add(name)
    s = wb.create_sheet(name)
    s["A1"] = t["name"]
    s["A1"].font = TITLE_FONT
    s["A1"].fill = GOLD_FILL
    s.merge_cells("A1:B1")
    s["A2"] = "Description:"
    s["A2"].font = BOLD
    s["B2"] = t.get("description", "")
    s["B2"].alignment = WRAP
    s["A3"] = "Page count:"
    s["A3"].font = BOLD
    s["B3"] = len(t["urls"])
    hr = 5
    s.cell(row=hr, column=1, value="#").fill = HEADER_FILL
    s.cell(row=hr, column=1).font = HEADER_FONT
    s.cell(row=hr, column=2, value="Page URL").fill = HEADER_FILL
    s.cell(row=hr, column=2).font = HEADER_FONT
    for i, url in enumerate(t["urls"], 1):
        rr = hr + i
        s.cell(row=rr, column=1, value=i).border = BORDER
        uc = s.cell(row=rr, column=2, value=url)
        uc.hyperlink = url
        uc.font = Font(color="0662BB", underline="single")
        uc.border = BORDER
    s.column_dimensions["A"].width = 6
    s.column_dimensions["B"].width = 95
    s.freeze_panes = "A6"

wb.save(OUT)
print(f"Saved: {OUT}")
print(f"Sheets: {len(wb.sheetnames)} -> Summary, All Pages, + {len(templates)} template sheets")
print(f"Total pages listed: {sum(len(t['urls']) for t in templates)}")
